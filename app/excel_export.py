"""Generate Excel (.xlsx) exports with Rupiah number formats.

Satu file cashflow lengkap ala spreadsheet acuan Leks Phone:
- Rekap Stok
- Penjualan
- Kas (cashflow)
- Operasional
- Ringkasan
"""
from __future__ import annotations

import io
from datetime import date, datetime
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

RP_FORMAT_NEG = '"Rp"#,##0;[Red]-"Rp"#,##0'

_header_font = Font(bold=True, color="FFFFFF")
_header_fill = PatternFill("solid", fgColor="1A2332")
_thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
_title_font = Font(bold=True, size=14)
_section_font = Font(bold=True, size=11)


def _style_header(ws: Worksheet, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin


def _auto_width(ws: Worksheet, min_w: int = 10, max_w: int = 40) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, length + 2))


def _money(ws: Worksheet, row: int, col: int, value) -> None:
    cell = ws.cell(row=row, column=col)
    try:
        num = float(value or 0)
    except (TypeError, ValueError):
        num = 0.0
    cell.value = num
    cell.number_format = RP_FORMAT_NEG
    cell.alignment = Alignment(horizontal="right")
    cell.border = _thin


def _text(ws: Worksheet, row: int, col: int, value) -> None:
    cell = ws.cell(row=row, column=col, value="" if value is None else str(value))
    cell.border = _thin
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def _date_cell(ws: Worksheet, row: int, col: int, value) -> None:
    cell = ws.cell(row=row, column=col)
    if isinstance(value, datetime):
        cell.value = value.date()
        cell.number_format = "DD/MM/YYYY"
    elif isinstance(value, date):
        cell.value = value
        cell.number_format = "DD/MM/YYYY"
    else:
        cell.value = str(value or "")
    cell.border = _thin


def _bool_sold(ws: Worksheet, row: int, col: int, sold: bool) -> None:
    cell = ws.cell(row=row, column=col, value="TRUE" if sold else "FALSE")
    cell.border = _thin
    cell.alignment = Alignment(horizontal="center")


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_BULAN = {
    1: "Januari",
    2: "Februari",
    3: "Maret",
    4: "April",
    5: "Mei",
    6: "Juni",
    7: "Juli",
    8: "Agustus",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Desember",
}


def _month_key_from_sale(row: dict) -> tuple[int, int] | None:
    """(year, month) dari sell_date atau field year/month."""
    if row.get("year") and row.get("month_num"):
        try:
            return int(row["year"]), int(row["month_num"])
        except (TypeError, ValueError):
            pass
    sd = row.get("sell_date")
    if isinstance(sd, datetime):
        return sd.year, sd.month
    if isinstance(sd, date):
        return sd.year, sd.month
    return None


def _month_key_from_ops(row: dict) -> tuple[int, int] | None:
    d = row.get("expense_date")
    if isinstance(d, datetime):
        return d.year, d.month
    if isinstance(d, date):
        return d.year, d.month
    return None


def _sheet_title_month(year: int, month: int) -> str:
    """Judul sheet Excel max 31 karakter — 'Februari 2026'."""
    name = f"{_BULAN.get(month, str(month))} {year}"
    return name[:31]


def _money_formula(ws: Worksheet, row: int, col: int, formula: str) -> None:
    """Cell angka dengan rumus Excel + format Rupiah."""
    cell = ws.cell(row=row, column=col, value=formula)
    cell.number_format = RP_FORMAT_NEG
    cell.alignment = Alignment(horizontal="right")
    cell.border = _thin
    cell.font = Font(bold=True)


def _write_sales_table(
    ws: Worksheet,
    sales: list[dict],
    *,
    start_row: int = 5,
    title: str = "",
    subtitle: str = "",
    exported_at: str = "",
    with_month_col: bool = False,
) -> dict:
    """
    Tulis tabel penjualan. Urutan kolom:
    No | Tanggal | [Bulan] | Metode | Pembeli | No. WhatsApp | Jenis Hp | IMEI | Serial |
    Harga beli | Charger | Harga jual | Biaya | Total Keuntungan | Catatan
    """
    if title:
        ws.cell(row=1, column=1, value=title).font = _title_font
        ws.cell(row=2, column=1, value=f"Diekspor: {exported_at}")
        if subtitle:
            ws.cell(row=3, column=1, value=subtitle)

    if with_month_col:
        cols = [
            "No",
            "Tanggal",
            "Bulan",
            "Metode Penjualan",
            "Pembeli",
            "No. WhatsApp",
            "Jenis Hp",
            "IMEI",
            "Serial",
            "Harga beli",
            "Harga Charger",
            "Harga jual",
            "Biaya lain-lain",
            "Total Keuntungan",
            "Catatan",
        ]
        # J=10 buy K=11 chg L=12 sell M=13 extra N=14 profit
        c_buy, c_chg, c_sell, c_extra, c_profit = 10, 11, 12, 13, 14
        c_label_total = 9  # Serial (sebelum harga beli)
        c_name = 7
    else:
        cols = [
            "No",
            "Tanggal",
            "Metode Penjualan",
            "Pembeli",
            "No. WhatsApp",
            "Jenis Hp",
            "IMEI",
            "Serial",
            "Harga beli",
            "Harga Charger",
            "Harga jual",
            "Biaya lain-lain",
            "Total Keuntungan",
            "Catatan",
        ]
        # I=9 buy J=10 chg K=11 sell L=12 extra M=13 profit
        c_buy, c_chg, c_sell, c_extra, c_profit = 9, 10, 11, 12, 13
        c_label_total = 8  # Serial
        c_name = 6

    for i, h in enumerate(cols, 1):
        ws.cell(row=start_row, column=i, value=h)
    _style_header(ws, start_row, len(cols))

    letters = {i: get_column_letter(i) for i in range(1, 16)}
    first_data = start_row + 1
    for idx, row in enumerate(sales, 1):
        r = start_row + idx
        buy = float(row.get("buy") or 0)
        charger = float(row.get("charger") or 0)
        sell = float(row.get("sell") or 0)
        extra = float(row.get("extra") or 0)
        buyer = (row.get("buyer") or "").strip() or "-"
        phone = (row.get("buyer_phone") or "").strip()
        _text(ws, r, 1, idx)
        _date_cell(ws, r, 2, row.get("sell_date"))
        if with_month_col:
            _text(ws, r, 3, row.get("month"))
            _text(ws, r, 4, row.get("method") or "—")
            _text(ws, r, 5, buyer)
            _text(ws, r, 6, phone)
            _text(ws, r, 7, row.get("name"))
            _text(ws, r, 8, row.get("imei"))
            _text(ws, r, 9, row.get("serial"))
            _money(ws, r, 10, buy)
            _money(ws, r, 11, charger)
            _money(ws, r, 12, sell)
            _money(ws, r, 13, extra)
            _money_formula(
                ws,
                r,
                14,
                f"={letters[12]}{r}-{letters[10]}{r}-{letters[11]}{r}-{letters[13]}{r}",
            )
            ws.cell(row=r, column=14).font = Font()
            _text(ws, r, 15, row.get("notes"))
        else:
            _text(ws, r, 3, row.get("method") or "—")
            _text(ws, r, 4, buyer)
            _text(ws, r, 5, phone)
            _text(ws, r, 6, row.get("name"))
            _text(ws, r, 7, row.get("imei"))
            _text(ws, r, 8, row.get("serial"))
            _money(ws, r, 9, buy)
            _money(ws, r, 10, charger)
            _money(ws, r, 11, sell)
            _money(ws, r, 12, extra)
            _money_formula(
                ws,
                r,
                13,
                f"={letters[11]}{r}-{letters[9]}{r}-{letters[10]}{r}-{letters[12]}{r}",
            )
            ws.cell(row=r, column=13).font = Font()
            _text(ws, r, 14, row.get("notes"))

    last_data = start_row + len(sales) if sales else start_row
    total_row = last_data + 1 if sales else start_row + 1

    if sales:
        # Hanya baris jumlah (rumus SUM) — tanpa label "Total"
        for col in (c_buy, c_chg, c_sell, c_extra, c_profit):
            L = letters[col]
            _money_formula(ws, total_row, col, f"=SUM({L}{first_data}:{L}{last_data})")
    else:
        for col in (c_buy, c_chg, c_sell, c_extra, c_profit):
            _money(ws, total_row, col, 0)

    return {
        "header_row": start_row,
        "first_data": first_data,
        "last_data": last_data,
        "total_row": total_row,
        "c_buy": c_buy,
        "c_chg": c_chg,
        "c_sell": c_sell,
        "c_extra": c_extra,
        "c_profit": c_profit,
        "n_sales": len(sales),
    }


def export_cashflow_workbook(
    *,
    stock_rows: list[dict],
    sales_rows: list[dict],
    cash_rows: list[dict],
    ops_rows: list[dict],
    summary: dict,
    partners_split: list[tuple],
) -> bytes:
    """
    stock_rows: purchase_date, name, supplier, buy, sold, notes, imei, serial, status, category
    sales_rows: sell_date, method, buyer, name, buy, charger, sell, extra, profit, notes, imei,
                month, year, month_num
    cash_rows / ops_rows / summary / partners_split: lihat pemanggil
    """
    wb = Workbook()
    exported_at = datetime.now().strftime("%d/%m/%Y %H:%M")

    # ---------- 1. Rekap Stok (seperti gid stok spreadsheet) ----------
    ws = wb.active
    ws.title = "Rekap Stok"
    ws["A1"] = "Leks Phone — Rekap Stok"
    ws["A1"].font = _title_font
    ws["A2"] = f"Diekspor: {exported_at}"
    ws["A3"] = "Mirip tab rekap stok spreadsheet (beli + status terjual)"

    cols = [
        "No",
        "Waktu Pembelian",
        "Barang",
        "Tempat Beli",
        "Harga beli",
        "Terjual",
        "Status",
        "Kategori",
        "IMEI",
        "Serial",
        "Catatan",
    ]
    hr = 5
    for i, h in enumerate(cols, 1):
        ws.cell(row=hr, column=i, value=h)
    _style_header(ws, hr, len(cols))

    for idx, it in enumerate(stock_rows, 1):
        r = hr + idx
        buy = float(it.get("buy") or 0)
        _text(ws, r, 1, idx)
        _date_cell(ws, r, 2, it.get("purchase_date"))
        _text(ws, r, 3, it.get("name"))
        _text(ws, r, 4, it.get("supplier"))
        _money(ws, r, 5, buy)
        _bool_sold(ws, r, 6, bool(it.get("sold")))
        _text(ws, r, 7, it.get("status"))
        _text(ws, r, 8, it.get("category"))
        _text(ws, r, 9, it.get("imei"))
        _text(ws, r, 10, it.get("serial"))
        _text(ws, r, 11, it.get("notes"))

    if stock_rows:
        first = hr + 1
        last = hr + len(stock_rows)
        tr = last + 1
        lab = ws.cell(row=tr, column=4, value="Total")
        lab.font = Font(bold=True)
        lab.alignment = Alignment(horizontal="right")
        lab.border = _thin
        _money_formula(ws, tr, 5, f"=SUM(E{first}:E{last})")
    _auto_width(ws)

    # ---------- 2. Penjualan (gabungan semua bulan) ----------
    ws2 = wb.create_sheet("Penjualan Semua")
    _write_sales_table(
        ws2,
        sales_rows,
        start_row=5,
        title="Leks Phone — Penjualan & Keuntungan (Semua Bulan)",
        subtitle="Gabungan semua penjualan — detail per bulan di sheet terpisah. Angka pakai rumus Excel.",
        exported_at=exported_at,
        with_month_col=True,
    )
    _auto_width(ws2)

    # ---------- 2b. Group by month ----------
    sales_by_month: dict[tuple[int, int], list[dict]] = {}
    for row in sales_rows:
        mk = _month_key_from_sale(row)
        if not mk:
            continue
        sales_by_month.setdefault(mk, []).append(row)

    ops_by_month: dict[tuple[int, int], list[dict]] = {}
    for row in ops_rows:
        mk = _month_key_from_ops(row)
        if not mk:
            continue
        ops_by_month.setdefault(mk, []).append(row)

    all_months = sorted(set(sales_by_month.keys()) | set(ops_by_month.keys()))

    # partner % from partners_split
    partner_pcts: list[tuple[str, float]] = []
    for t in partners_split or []:
        if len(t) >= 2:
            partner_pcts.append((str(t[0]), float(t[1] or 0)))

    monthly_stats: list[dict] = []
    for y, m in all_months:
        srows = sales_by_month.get((y, m), [])
        orows = ops_by_month.get((y, m), [])
        omzet = sum(float(x.get("sell") or 0) for x in srows)
        laba = 0.0
        for x in srows:
            p = x.get("profit")
            if p is None:
                p = (
                    float(x.get("sell") or 0)
                    - float(x.get("buy") or 0)
                    - float(x.get("charger") or 0)
                    - float(x.get("extra") or 0)
                )
            laba += float(p or 0)
        ops_m = sum(float(x.get("amount") or 0) for x in orows)
        monthly_stats.append(
            {
                "year": y,
                "month": m,
                "label": f"{_BULAN.get(m, m)} {y}",
                "units": len(srows),
                "omzet": omzet,
                "laba": laba,
                "ops": ops_m,
                "laba_bersih": laba - ops_m,
                "sales": srows,
                "ops_rows": orows,
            }
        )

    # ---------- 2c. Laporan Bulanan (rekap perbandingan) ----------
    ws_m = wb.create_sheet("Laporan Bulanan")
    ws_m["A1"] = "Leks Phone — Laporan per Bulan"
    ws_m["A1"].font = _title_font
    ws_m["A2"] = f"Diekspor: {exported_at}"
    ws_m["A3"] = "Ringkasan omzet, laba, operasional tiap bulan (seperti tab bulanan di spreadsheet)"

    mcols = [
        "No",
        "Bulan",
        "Unit terjual",
        "Omzet (Rp)",
        "Laba unit (Rp)",
        "Operasional (Rp)",
        "Laba bersih (Rp)",
    ]
    # tambah kolom partner
    for name, _pct in partner_pcts:
        mcols.append(f"Bagi {name} (Rp)")

    hr_m = 5
    for i, h in enumerate(mcols, 1):
        ws_m.cell(row=hr_m, column=i, value=h)
    _style_header(ws_m, hr_m, len(mcols))

    first_m = hr_m + 1
    for idx, st in enumerate(monthly_stats, 1):
        r = hr_m + idx
        _text(ws_m, r, 1, idx)
        _text(ws_m, r, 2, st["label"])
        # unit & angka disimpan dulu (nilai); total baris pakai SUM
        cell_u = ws_m.cell(row=r, column=3, value=int(st["units"]))
        cell_u.border = _thin
        _money(ws_m, r, 4, st["omzet"])
        _money(ws_m, r, 5, st["laba"])
        _money(ws_m, r, 6, st["ops"])
        # laba bersih = laba − ops (rumus)
        _money_formula(ws_m, r, 7, f"=E{r}-F{r}")
        ws_m.cell(row=r, column=7).font = Font()
        for pi, (pname, pct) in enumerate(partner_pcts):
            # bagi hasil = laba * pct/100
            _money_formula(ws_m, r, 8 + pi, f"=E{r}*{pct}/100")
            ws_m.cell(row=r, column=8 + pi).font = Font()

    if monthly_stats:
        last_m = hr_m + len(monthly_stats)
        tr = last_m + 1
        _text(ws_m, tr, 2, "TOTAL")
        ws_m.cell(row=tr, column=2).font = Font(bold=True)
        _money_formula(ws_m, tr, 3, f"=SUM(C{first_m}:C{last_m})")
        for col in range(4, 8 + len(partner_pcts)):
            L = get_column_letter(col)
            _money_formula(ws_m, tr, col, f"=SUM({L}{first_m}:{L}{last_m})")
    _auto_width(ws_m)

    # ---------- 2d. Sheet per bulan (ala tab spreadsheet + rumus) ----------
    used_titles: set[str] = set()
    for st in monthly_stats:
        y, m = st["year"], st["month"]
        title = _sheet_title_month(y, m)
        base = title
        n = 2
        while title in used_titles or title in [s.title for s in wb.worksheets]:
            title = f"{base[:28]} {n}"[:31]
            n += 1
        used_titles.add(title)

        wsm = wb.create_sheet(title)
        meta = _write_sales_table(
            wsm,
            st["sales"],
            start_row=5,
            title=f"Leks Phone — Penjualan {st['label']}",
            subtitle="Catatan = rincian biaya lain-lain transaksi. Total & ringkasan pakai rumus Excel.",
            exported_at=exported_at,
            with_month_col=False,
        )
        # Layout bawah seperti screenshot:
        #   F: Unit terjual | G: n     | I: Laba unit   | J: =total untung
        #   F: Omzet        | H: =jual | I: Laba bersih | J: =laba−ops
        #                    |         | I: BAGI HASIL  |
        #                    |         | I: Adhit 50%   | J: =…
        tr = meta["total_row"]
        c_buy = meta["c_buy"]  # F
        c_chg = meta["c_chg"]  # G
        c_sell = meta["c_sell"]  # H
        c_extra = meta["c_extra"]  # I
        c_profit = meta["c_profit"]  # J
        Lf = get_column_letter(c_buy)
        Lg = get_column_letter(c_chg)
        Lh = get_column_letter(c_sell)
        Li = get_column_letter(c_extra)
        Lj = get_column_letter(c_profit)

        def _lab(ws, row, col, text, bold=True):
            cell = ws.cell(row=row, column=col, value=text)
            cell.font = Font(bold=bold)
            cell.alignment = Alignment(horizontal="right")
            cell.border = _thin
            return cell

        r = tr + 2  # satu baris kosong setelah Total (seperti screenshot)

        # Baris 1: Unit terjual | n || Laba unit | formula
        _lab(wsm, r, c_buy, "Unit terjual")
        cell_u = wsm.cell(row=r, column=c_chg, value=meta["n_sales"])
        cell_u.font = Font(bold=True)
        cell_u.alignment = Alignment(horizontal="right")
        cell_u.border = _thin
        _lab(wsm, r, c_extra, "Laba unit")
        _money_formula(wsm, r, c_profit, f"={Lj}{tr}")
        r_laba = r
        r += 1

        # Baris 2: Omzet | (value di Harga jual) || Laba bersih | formula
        _lab(wsm, r, c_buy, "Omzet")
        _money_formula(wsm, r, c_sell, f"={Lh}{tr}")
        _lab(wsm, r, c_extra, "Laba bersih")
        r_bersih = r
        r += 1

        # Ops tersembunyi di sel untuk rumus (jika ada ops, isi nanti; default 0)
        # taruh di kolom G baris r_bersih sementara sebagai helper? Lebih bersih: cell ops di G baris unit
        # Simpan r_ops_val di kolom charger baris Laba bersih (tersembunyi visually as 0)
        # Actually use a dedicated cell after bagi hasil for ops total, or put ops amount in G of r_bersih unused
        r_ops_cell_row = r_bersih
        r_ops_cell_col = c_chg  # G next to Omzet label area unused on row 2 left... wait row2 G empty
        # Put ops amount at G of r_bersih with formula/value, label not shown - for laba bersih formula only
        # User screenshot doesn't show Operasional — keep G empty for row1 value only

        # BAGI HASIL block (kanan, kolom I–J)
        r += 1
        _lab(wsm, r, c_extra, "BAGI HASIL")
        r += 1
        for pname, pct in partner_pcts:
            _lab(wsm, r, c_extra, f"{pname} ({pct:.0f}%)")
            _money_formula(wsm, r, c_profit, f"={Lj}{r_laba}*{pct}/100")
            r += 1

        # Operasional: hitung lewat rumus jika ada detail di bawah; tautkan Laba bersih
        ops_formula_cell = None
        if st["ops_rows"]:
            r += 2
            wsm.cell(row=r, column=1, value="BIAYA OPERASIONAL BULAN INI").font = _section_font
            r += 1
            for i, h in enumerate(["No", "Tanggal", "Kategori", "Jumlah (Rp)", "Keterangan"], 1):
                wsm.cell(row=r, column=i, value=h)
            _style_header(wsm, r, 5)
            r += 1
            ops_first = r
            for i, o in enumerate(st["ops_rows"], 1):
                _text(wsm, r, 1, i)
                _date_cell(wsm, r, 2, o.get("expense_date"))
                _text(wsm, r, 3, o.get("category"))
                _money(wsm, r, 4, o.get("amount"))
                _text(wsm, r, 5, o.get("description"))
                r += 1
            ops_last = r - 1
            _lab(wsm, r, 3, "Total")
            _money_formula(wsm, r, 4, f"=SUM(D{ops_first}:D{ops_last})")
            ops_formula_cell = f"D{r}"
            # Laba bersih = Laba unit − Total ops
            _money_formula(wsm, r_bersih, c_profit, f"={Lj}{r_laba}-{ops_formula_cell}")
        else:
            # Laba bersih = Laba unit (ops 0)
            _money_formula(wsm, r_bersih, c_profit, f"={Lj}{r_laba}")

        _auto_width(wsm)

    # ---------- 3. Kas cashflow ----------
    ws3 = wb.create_sheet("Kas")
    ws3["A1"] = "Leks Phone — Buku Kas (Cashflow)"
    ws3["A1"].font = _title_font
    ws3["A2"] = f"Diekspor: {exported_at}"
    ws3["A3"] = "Semua transaksi masuk / keluar"

    cols3 = ["No", "Tanggal", "Arah", "Jenis", "Jumlah (Rp)", "Keterangan", "Oleh"]
    hr3 = 5
    for i, h in enumerate(cols3, 1):
        ws3.cell(row=hr3, column=i, value=h)
    _style_header(ws3, hr3, len(cols3))

    for idx, e in enumerate(cash_rows, 1):
        r = hr3 + idx
        amt = float(e.get("amount") or 0)
        direction = e.get("direction") or ""
        signed = amt if direction == "in" else -amt
        _text(ws3, r, 1, idx)
        _date_cell(ws3, r, 2, e.get("txn_date"))
        _text(ws3, r, 3, "Masuk" if direction == "in" else "Keluar")
        _text(ws3, r, 4, e.get("entry_type"))
        _money(ws3, r, 5, signed)
        _text(ws3, r, 6, e.get("description"))
        _text(ws3, r, 7, e.get("created_by"))

    if cash_rows:
        first_c = hr3 + 1
        last_c = hr3 + len(cash_rows)
        tr = last_c + 1
        # Total masuk = SUMIF Masuk; keluar = -SUMIF Keluar; saldo = SUM signed
        _text(ws3, tr, 4, "Total masuk")
        ws3.cell(row=tr, column=4).font = Font(bold=True)
        _money_formula(
            ws3,
            tr,
            5,
            f'=SUMIF(C{first_c}:C{last_c},"Masuk",E{first_c}:E{last_c})',
        )
        _text(ws3, tr + 1, 4, "Total keluar")
        ws3.cell(row=tr + 1, column=4).font = Font(bold=True)
        # keluar tersimpan negatif → ambil abs
        _money_formula(
            ws3,
            tr + 1,
            5,
            f'=ABS(SUMIF(C{first_c}:C{last_c},"Keluar",E{first_c}:E{last_c}))',
        )
        _text(ws3, tr + 2, 4, "Saldo buku kas")
        ws3.cell(row=tr + 2, column=4).font = Font(bold=True)
        _money_formula(ws3, tr + 2, 5, f"=SUM(E{first_c}:E{last_c})")
    _auto_width(ws3)

    # ---------- 4. Operasional ----------
    ws4 = wb.create_sheet("Operasional")
    ws4["A1"] = "Leks Phone — Biaya Operasional"
    ws4["A1"].font = _title_font
    ws4["A2"] = f"Diekspor: {exported_at}"

    cols4 = ["No", "Tanggal", "Kategori", "Jumlah (Rp)", "Keterangan"]
    hr4 = 4
    for i, h in enumerate(cols4, 1):
        ws4.cell(row=hr4, column=i, value=h)
    _style_header(ws4, hr4, len(cols4))

    for idx, o in enumerate(ops_rows, 1):
        r = hr4 + idx
        amt = float(o.get("amount") or 0)
        _text(ws4, r, 1, idx)
        _date_cell(ws4, r, 2, o.get("expense_date"))
        _text(ws4, r, 3, o.get("category"))
        _money(ws4, r, 4, amt)
        _text(ws4, r, 5, o.get("description"))
    if ops_rows:
        first_o = hr4 + 1
        last_o = hr4 + len(ops_rows)
        tr = last_o + 1
        lab = ws4.cell(row=tr, column=3, value="Total")
        lab.font = Font(bold=True)
        lab.alignment = Alignment(horizontal="right")
        lab.border = _thin
        _money_formula(ws4, tr, 4, f"=SUM(D{first_o}:D{last_o})")
    _auto_width(ws4)

    # ---------- 5. Ringkasan cashflow ----------
    ws5 = wb.create_sheet("Ringkasan")
    ws5["A1"] = "Leks Phone — Ringkasan Cashflow"
    ws5["A1"].font = _title_font
    ws5["A2"] = f"Diekspor: {exported_at}"

    ws5["A4"] = "POSISI KEUANGAN"
    ws5["A4"].font = _section_font
    metrics = [
        ("Modal disetor", summary.get("modal_setor")),
        ("Modal di barang (stok ready)", summary.get("modal_barang")),
        ("Kas di rekening (modal − stok)", summary.get("kas_rekening")),
        ("Saldo buku kas (cashflow penuh)", summary.get("saldo_buku")),
        ("Unit ready (qty item)", summary.get("ready_count")),
        ("Unit terjual (baris)", summary.get("sold_count")),
    ]
    r = 5
    for label, val in metrics:
        _text(ws5, r, 1, label)
        if "Unit" in label or "qty" in label.lower():
            _text(ws5, r, 2, val)
        else:
            _money(ws5, r, 2, val)
        r += 1

    r += 1
    ws5.cell(row=r, column=1, value="HASIL USAHA (SEMUA PENJUALAN)").font = _section_font
    r += 1
    for label, val in [
        ("Total omzet (harga jual)", summary.get("total_omzet")),
        ("Total laba unit", summary.get("total_laba")),
        ("Total operasional", summary.get("total_ops")),
        ("Laba bersih (laba unit − ops)", summary.get("laba_bersih")),
    ]:
        _text(ws5, r, 1, label)
        _money(ws5, r, 2, val)
        r += 1

    r += 1
    ws5.cell(row=r, column=1, value="BAGI HASIL (dari total laba unit)").font = _section_font
    r += 1
    for name, pct, amt in partners_split or []:
        _text(ws5, r, 1, f"{name} ({pct:.0f}%)")
        _money(ws5, r, 2, amt)
        r += 1

    r += 2
    ws5.cell(row=r, column=1, value="Catatan").font = Font(bold=True)
    r += 1
    ws5.cell(
        row=r,
        column=1,
        value=(
            "File setara spreadsheet acuan: Rekap Stok + Penjualan Semua + Laporan Bulanan "
            "+ sheet tiap bulan (Januari/Februari/…) + Kas + Operasional. "
            "Kolom uang berformat Rupiah (bisa dijumlah di Excel). "
            "Sumber data: database Leks Phone."
        ),
    )
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws5.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws5.row_dimensions[r].height = 45
    _auto_width(ws5, max_w=48)

    return workbook_to_bytes(wb)


# --- legacy helpers (masih bisa dipanggil jika perlu) ---
def export_laporan(**kwargs) -> bytes:
    """Deprecated: gunakan export_cashflow_workbook."""
    return export_cashflow_workbook(
        stock_rows=[],
        sales_rows=kwargs.get("sold_rows") or [],
        cash_rows=[],
        ops_rows=[],
        summary={
            "modal_setor": 0,
            "modal_barang": 0,
            "kas_rekening": 0,
            "saldo_buku": 0,
            "ready_count": 0,
            "sold_count": len(kwargs.get("sold_rows") or []),
            "total_omzet": kwargs.get("omzet") or 0,
            "total_laba": kwargs.get("laba") or 0,
            "total_ops": kwargs.get("ops") or 0,
            "laba_bersih": kwargs.get("laba_bersih") or 0,
        },
        partners_split=kwargs.get("bagi_hasil") or [],
    )


def export_stok(items: Iterable[dict]) -> bytes:
    rows = []
    for it in items:
        rows.append(
            {
                "purchase_date": it.get("purchase_date"),
                "name": it.get("name"),
                "supplier": it.get("supplier"),
                "buy": it.get("buy"),
                "sold": (it.get("status") == "sold") or (it.get("qty_rem", 1) <= 0),
                "status": it.get("status"),
                "category": it.get("category"),
                "imei": it.get("imei"),
                "serial": it.get("serial"),
                "notes": "",
            }
        )
    return export_cashflow_workbook(
        stock_rows=rows,
        sales_rows=[],
        cash_rows=[],
        ops_rows=[],
        summary={
            "modal_setor": 0,
            "modal_barang": 0,
            "kas_rekening": 0,
            "saldo_buku": 0,
            "ready_count": 0,
            "sold_count": 0,
            "total_omzet": 0,
            "total_laba": 0,
            "total_ops": 0,
            "laba_bersih": 0,
        },
        partners_split=[],
    )


def export_kas(entries: Iterable[dict], saldo: float) -> bytes:
    return export_cashflow_workbook(
        stock_rows=[],
        sales_rows=[],
        cash_rows=list(entries),
        ops_rows=[],
        summary={
            "modal_setor": 0,
            "modal_barang": 0,
            "kas_rekening": 0,
            "saldo_buku": saldo,
            "ready_count": 0,
            "sold_count": 0,
            "total_omzet": 0,
            "total_laba": 0,
            "total_ops": 0,
            "laba_bersih": 0,
        },
        partners_split=[],
    )
